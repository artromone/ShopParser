import os
import json
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def adjust_column_widths(sheet):
    for column in sheet.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        column_letter = get_column_letter(column[0].column)
        adjusted_width = (max_length + 2) * 1.2
        sheet.column_dimensions[column_letter].width = adjusted_width


def format_datetime(datetime_str):
    datetime_obj = datetime.datetime.strptime(datetime_str, "%Y-%m-%dT%H:%M:%S")
    formatted_datetime = datetime_obj.strftime("%d.%m.%y-%H:%M")
    return formatted_datetime


def replace_shop_name(shop_name):
    replace_file_path = os.path.join("sys", "shopnamereplace.json")

    # Load data from shopnamereplace.json for shop name replacements
    with open(replace_file_path, "r", encoding="utf-8") as replace_file:
        replace_data = json.load(replace_file)

    # Replace shop name if there is a matching entry in shopnamereplace.json
    for entry in replace_data:
        if entry.get("original_name") == shop_name:
            return entry.get("short_name")

    return shop_name


def create_export_file(data):
    # Path to the output directory
    output_dir = "output/"

    # Load excluded shop names
    excluded_file_path = os.path.join("sys", "excluded.json")
    excluded_shops = []
    if os.path.exists(excluded_file_path):
        with open(excluded_file_path, "r", encoding="utf-8") as excluded_file:
            excluded_shops = json.load(excluded_file)

    # Create a new Excel workbook for export
    export_workbook = Workbook()

    # Remove the default "Sheet" that is automatically created
    default_sheet = export_workbook["Sheet"]
    export_workbook.remove(default_sheet)

    # Check if there is any data available
    if not data:
        # Create a single sheet named "No data" if there is no data
        export_sheet = export_workbook.active
        export_sheet.title = "No data"
    else:
        # Group the data by month and day
        data_by_month = {}
        for item in data:
            receipt = item.get("ticket", {}).get("document", {}).get("receipt", {})
            datetime_str = receipt.get("dateTime")
            shop_name = replace_shop_name(receipt.get("user"))
            if datetime_str and shop_name not in excluded_shops:
                datetime_obj = datetime.datetime.strptime(datetime_str, "%Y-%m-%dT%H:%M:%S")
                month = datetime_obj.strftime("%Y-%m")
                day = datetime_obj.date()
                if month not in data_by_month:
                    data_by_month[month] = {}
                if day not in data_by_month[month]:
                    data_by_month[month][day] = []
                data_by_month[month][day].append(item)

        # Create separate sheets for each month and add data for each day
        for month, month_data in data_by_month.items():
            month_sheet = export_workbook.create_sheet(title=month)

            for day, day_data in month_data.items():
                export_sheet = month_sheet

                # Add a separator row if it's not the first day
                if day != min(month_data.keys()):
                    export_sheet.append([])

                # Add a header row
                export_sheet.append(["Date/Time", "Shop", "Product", "Price", "Quantity", "Sum"])
                export_sheet.freeze_panes = "A2"

                # Apply bold font to the header row
                export_bold_font = Font(bold=True)
                for cell in export_sheet[1]:
                    cell.font = export_bold_font

                # Export data to the export sheet
                sum_total = 0
                for item in day_data:
                    receipt = item.get("ticket", {}).get("document", {}).get("receipt", {})
                    shop_name = replace_shop_name(receipt.get("user"))
                    for product in receipt.get("items", []):
                        row = [
                            format_datetime(receipt.get("dateTime")), shop_name, product.get("name"),
                            product.get("price") / 100, product.get("quantity"), product.get("sum") / 100
                        ]
                        export_sheet.append(row)
                        sum_total += product.get("sum") / 100

                # Add a row with the total sum for the day
                export_sheet.append([])  # Add an empty row
                total_row = ["Total", "", "", "", "", sum_total]
                export_sheet.append(total_row)

                # Apply bold font to the total row
                for row in export_sheet.iter_rows(max_row=export_sheet.max_row, max_col=export_sheet.max_column):
                    for cell in row:
                        if cell.row == export_sheet.max_row:
                            cell.font = export_bold_font

            # Adjust column widths for the month sheet
            adjust_column_widths(month_sheet)

    # Save the export workbook as export.xlsx
    export_file = os.path.join(output_dir, "export.xlsx")
    export_workbook.save(export_file)

    print(f"Export file {export_file} created successfully.")


# Path to the input directory
input_dir = "input/"

# Get a list of JSON files in the directory
json_files = [file for file in os.listdir(input_dir) if file.endswith(".json")]

data = []  # Store the data from all JSON files

for json_file in json_files:
    # Load data from the JSON file
    with open(os.path.join(input_dir, json_file), "r", encoding="utf-8") as file:
        data.extend(json.load(file))

# Create the export file with separate sheets for each month and add separators between days
create_export_file(data)
